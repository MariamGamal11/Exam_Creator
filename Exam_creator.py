from tkinter import *
from tkinter import messagebox, ttk
from random import randint, choice
import json
import random
from tkinter.ttk import Treeview, Style
from PIL import Image, ImageTk
import openpyxl
from openpyxl import load_workbook


###########import json files

with open (r"JSON_files/quiz1_questions.json", 'r', encoding='utf-8') as file1:
    quiz1 = json.load(file1)
with open (r"JSON_files/quiz1_data.json", 'r') as data1:
    data1 = json.load(data1)
with open (r"JSON_files/quiz2_questions.json", 'r') as file2:
    quiz2 = json.load(file2)
with open (r"JSON_files/quiz2_data.json", 'r') as data2:
    data2 = json.load(data2)
with open (r"JSON_files/quiz3_questions.json", 'r') as file3:
    quiz3 = json.load(file3)
with open (r"JSON_files/quiz3_data.json", 'r') as data3:
    data3 = json.load(data3)
with open (r"JSON_files/quiz4_questions.json", 'r') as file4:
    quiz4 = json.load(file4)
with open (r"JSON_files/quiz4_data.json", 'r') as data4:
    data4 = json.load(data4)
with open (r"JSON_files/quiz5_questions.json", 'r') as file5:
    quiz5 = json.load(file5)
with open (r"JSON_files/quiz5_data.json", 'r') as data5:
    data5 = json.load(data5)
with open (r"JSON_files/quiz6_questions.json", 'r') as file6:
    quiz6 = json.load(file6)
with open (r"JSON_files/quiz6_data.json", 'r') as data6:
    data6 = json.load(data6)
with open (r"JSON_files/quiz7_questions.json", 'r') as file7:
    quiz7 = json.load(file7)
with open (r"JSON_files/quiz7_data.json", 'r') as data7:
    data7 = json.load(data7)
with open (r"JSON_files/quiz8_questions.json", 'r') as file8:
    quiz8 = json.load(file8)
with open (r"JSON_files/quiz8_data.json", 'r') as data8:
    data8 = json.load(data8)
with open (r"JSON_files/quiz9_questions.json", 'r') as file9:
    quiz9 = json.load(file9)
with open (r"JSON_files/quiz9_data.json", 'r') as data9:
    data9 = json.load(data9)
with open(r"JSON_files/package.json", "r", encoding="utf-8") as file:
    quiz_scores = json.load(file)
## store Files, Paths, Data
quizzes_list = [quiz1,quiz2,quiz3,quiz4,quiz5,quiz6,quiz7,quiz8,quiz9]
quizzes_file = ["quiz1_questions.json",
                "quiz2_questions.json",
                "quiz3_questions.json",
                "quiz5_questions.json",
                "quiz6_questions.json",
                "quiz7_questions.json",
                "quiz8_questions.json",
                "quiz9_questions.json",]
data_list = [data1,data2,data3,data4,data5,data6,data7,data8,data9]


############### Results Page ##################

def Result_page():
    percentage = []
    def calculate_percentage(score, max_score):
        if score == "Not attempted":
            return "Not attempted"
        percentage.append((float(score) / max_score) * 100)
        return f"{(float(score) / max_score) * 100:.2f}%"

    def grade(score, total_score):
        if score == "Not attempted":
            return "N/A"
        score = int(score)
        if (score /total_score)*100 >= 90:
            return "A"
        elif (score /total_score)*100 >= 80:
            return "B"
        elif (score /total_score)*100 >= 70:
            return "C"
        elif (score /total_score)*100 >= 60:
            return "D"
        else:
            return "F"

    def get_feedback(grade1):
        feedbacks = {
            'A+': 'Excellent performance!',
            'A': 'Great job!',
            'B': 'Good effort,\nbut room for \nimprovement.',
            'C': 'Needs improvement',
            'D': 'Passed,\nbut needs\nmore work',
            'F': 'Failed,\ntry harder \nnext time',
            'N/A': 'No attempt made'
        }
        return feedbacks.get(grade1, 'No feedback available.')

    ##########################################Window setup##########################################
    global results_window
    results_window = Tk()
    results_window.geometry('1920x1080')
    results_window.configure(background='#003154')
    results_window.title('Results')
    results_window.config(padx=5, pady=5)
    #label

    welcome=Label(results_window,
                text="Welcome to your results".upper(),
                background = '#003154',
                font = " calibre 40 bold",
                fg = "white",

                compound = "center"
    )
    welcome.place(x=500, y=0)
    ###################################image#################
    image1=Image.open(r'Images\replay.png')
    image_resize=image1.resize((90,90))
    image_retake=ImageTk.PhotoImage(image_resize)
    ##########################################Frames##########################################
    parent1 = Frame(results_window, background="red", width=1800, height=800)
    parent1.grid_propagate(False)
    parent1.grid(row=0, column=0, padx=50, pady=100)

    ##########################################Treeview setup##########################################
    trv = Treeview(parent1, columns=(1, 2, 3, 4, 5, 6, 7, 8, 9, 10), show='headings', height=50)
    trv.grid(row=1, column=0, columnspan=1)

    ##########################################Headings##########################################
    trv.heading(1, text='Category')
    for i in range(2, 11):
        trv.heading(i, text=f'QUIZ {i-1}')

    ##########################################Columns##########################################
    for i in range(1, 11):
        trv.column(i, width=180, anchor='center')

    trv.tag_configure('bold', font=('Arial', 14, 'bold'))
    trv.tag_configure('bold_Feedback', font=('Arial', 12, 'bold'))

    ##########################################Results##########################################
    # First Attempt
    
    first = []
    for i in range(9):
        if len(quiz_scores['scores'][i]) > 0:
            first.append(quiz_scores['scores'][i][0])
        else:
            first.append("Not Attempted")
    trv.insert('', 'end', values=(
        "First Attempt",
        first[0],
        first[1],
        first[2],
        first[3],
        first[4],
        first[5],
        first[6],
        first[7],
        first[8]

    ), tags=('bold',))

    # Second attempt
    second = []
    for i in range(9):
        if len(quiz_scores['scores'][i]) > 1:
            second.append(quiz_scores['scores'][i][1])
        else:
            second.append("Not Attempted")
    trv.insert('', 'end', values=(
        "Second attempt ",
        second[0],
        second[1],
        second[2],
        second[3],
        second[4],
        second[5],
        second[6],
        second[7],
        second[8]

    ), tags=('bold',))
    
    #third attempt
    
    third = []
    for i in range(9):
        if len(quiz_scores['scores'][i]) > 2:
            third.append(quiz_scores['scores'][i][2])
        else:
            third.append("Not Attempted")
    trv.insert('', 'end', values=(
        "Third Attempt",
        third[0],
        third[1],
        third[2],
        third[3],
        third[4],
        third[5],
        third[6],
        third[7],
        third[8]
    ), tags=('bold',))

    ### max grades in each quiz

    try:
        quiz_1_score = max(quiz_scores['scores'][0])
    except:
        quiz_1_score = "Not attempted"
    try:
        quiz_2_score = max(quiz_scores['scores'][1])
    except:
        quiz_2_score = "Not attempted"
    try:
        quiz_3_score = max(quiz_scores['scores'][2])
    except:
        quiz_3_score = "Not attempted"
    try:
        quiz_4_score = max(quiz_scores['scores'][3])
    except:
        quiz_4_score = "Not attempted"
    try:
        quiz_5_score = max(quiz_scores['scores'][4])
    except:
        quiz_5_score = "Not attempted"
    try:
        quiz_6_score = max(quiz_scores['scores'][5])
    except:
        quiz_6_score = "Not attempted"
    try:
        quiz_7_score = max(quiz_scores['scores'][6])
    except:
        quiz_7_score = "Not attempted"
    try:
        quiz_8_score = max(quiz_scores['scores'][7])
    except:
        quiz_8_score = "Not attempted"
    try:
        quiz_9_score = max(quiz_scores['scores'][9])
    except:
        quiz_9_score = "Not attempted"


    # Percentage
    
    trv.insert('', 'end', values=(
        "Percentage",
        calculate_percentage(quiz_1_score, (data1["Grade per Question"] * data1["Number of Questions"])),
        calculate_percentage(quiz_2_score, (data2["Grade per Question"] * data2["Number of Questions"])),
        calculate_percentage(quiz_3_score, (data3["Grade per Question"] * data3["Number of Questions"])),
        calculate_percentage(quiz_4_score, (data4["Grade per Question"] * data4["Number of Questions"])),
        calculate_percentage(quiz_5_score, (data5["Grade per Question"] * data5["Number of Questions"])),
        calculate_percentage(quiz_6_score, (data6["Grade per Question"] * data6["Number of Questions"])),
        calculate_percentage(quiz_7_score, (data7["Grade per Question"] * data7["Number of Questions"])),
        calculate_percentage(quiz_8_score, (data8["Grade per Question"] * data8["Number of Questions"])),
        calculate_percentage(quiz_9_score, (data9["Grade per Question"] * data9["Number of Questions"]))
    ), tags=('bold',))

    # letter Grade
    
    trv.insert('', 'end', values=(
        "Grade",
        grade(quiz_1_score, (data1["Grade per Question"] * data1["Number of Questions"])),
        grade(quiz_2_score, (data2["Grade per Question"] * data2["Number of Questions"])),
        grade(quiz_3_score, (data3["Grade per Question"] * data3["Number of Questions"])),
        grade(quiz_4_score, (data4["Grade per Question"] * data4["Number of Questions"])),
        grade(quiz_5_score, (data5["Grade per Question"] * data5["Number of Questions"])),
        grade(quiz_6_score, (data6["Grade per Question"] * data6["Number of Questions"])),
        grade(quiz_7_score, (data7["Grade per Question"] * data7["Number of Questions"])),
        grade(quiz_8_score, (data8["Grade per Question"] * data8["Number of Questions"])),
        grade(quiz_9_score, (data9["Grade per Question"] * data9["Number of Questions"]))
    ), tags=('bold',))

    ########################## Feedback ##############################
    # Feedback (Using the grade function to get feedback)
    
    trv.insert('', 'end', values=(
        "Feedback",
        get_feedback(grade(quiz_1_score, (data1["Grade per Question"] * data1["Number of Questions"]))),
        get_feedback(grade(quiz_2_score, (data2["Grade per Question"] * data2["Number of Questions"]))),
        get_feedback(grade(quiz_3_score, (data3["Grade per Question"] * data3["Number of Questions"]))),
        get_feedback(grade(quiz_4_score, (data4["Grade per Question"] * data4["Number of Questions"]))),
        get_feedback(grade(quiz_5_score, (data5["Grade per Question"] * data5["Number of Questions"]))),
        get_feedback(grade(quiz_6_score, (data6["Grade per Question"] * data6["Number of Questions"]))),
        get_feedback(grade(quiz_7_score, (data7["Grade per Question"] * data7["Number of Questions"]))),
        get_feedback(grade(quiz_8_score, (data8["Grade per Question"] * data8["Number of Questions"]))),
        get_feedback(grade(quiz_9_score, (data9["Grade per Question"] * data9["Number of Questions"]))),
    ), tags=('bold_Feedback',))

    retake_button = Button(results_window,image=image_retake,width=140,height=90,background='#003154'
                        ,activebackground='#003154',activeforeground='red'
                        ,bd=0,text = 'Retake the exam',
                        compound = 'center', fg = "red",
                        font = ("Arial", 12, "italic", "bold"), command= choose_Exam)
    retake_button.place(x=1350, y=0)
    # Style
    style = Style()
    style.configure("Treeview.Heading", font=("Arial bold", 14))
    style.configure("Treeview", font=("Arial", 11), background="#D3D3D3", foreground="black",
                    rowheight=130, fieldbackground="#D3D3D3")

    # Run
    results_window.mainloop()






##################################Take Exam Page##############################

def TakeExam():

    # Sumbit Exam button functionality

    def submit_exam():
        global result, avg_grade, l
        result = 0

        # Check Questions (error if there is an unsolved questions)
        for m in range(data_list[chosen_quiz]['Number of Questions']):
            if not intvar_list[m].get():
                messagebox.showinfo("Error", "all questions are required")
                return
            
            #calculate total grade
            
            if(intvar_list[m].get() == ((m*4) + quizzes_list[chosen_quiz][random_numbers_list[m]]['answer'])):
                result+= data_list[chosen_quiz]['Grade per Question']

        quiz_scores['scores'][chosen_quiz].append(result)
        # Update JSON File
        with open(r"JSON_files/package.json", "w") as file:
            json.dump(quiz_scores, file, indent=1)
        try:
            avg_grade = sum(quizzes_list[chosen_quiz][0]['attempts_grades'])/len(quizzes_list[chosen_quiz][0]['attempts_grades'])
        except:
            avg_grade = 0
        Take_Exam.destroy()
        Result_page()

    #create a Take Exam window
    Take_Exam = Tk()
    Take_Exam.geometry("600x700")
    Take_Exam.title('window')

    # Canvas
    canvas = Canvas(Take_Exam, bg="#003154", scrollregion=(0, 0, 2000, data_list[chosen_quiz]['Number of Questions'] * 200))
    canvas.pack(expand=True, fill="both")


    # Scrollbar
    scroll_y = Scrollbar(Take_Exam, orient='vertical', command=canvas.yview)
    canvas.configure(yscrollcommand=scroll_y.set)
    scroll_y.place(relx=1, rely=0, relheight=1, anchor='ne')


    # Container frame inside the canvas (to include it in the scroll)
    container = Frame(canvas, bg="#003154", width=580)
    canvas.create_window((0, 0), window=container, anchor='nw', width= 580)

    canvas.bind('<MouseWheel>', lambda event: canvas.yview_scroll(-event.delta//60), "units")

    # intvar list
    intvar_list = []
    for k in range(data_list[chosen_quiz]['Number of Questions']):
        intvar_list.append(IntVar())
    y = 1

    # Question generator

    random_numbers_list = []

    for i in range(data_list[chosen_quiz]['Number of Questions']):
        
        # Questions Randomizer
        
        random_number = randint(1, len(quizzes_list[chosen_quiz])-1) # Question bank random questions generator
        while(random_number in random_numbers_list): ### Ensure that there are no repeated Questions
            random_number = randint(1, len(quizzes_list[chosen_quiz])-1)
        random_numbers_list.append(random_number) ## Store taken Quizzes


        # question container
        question = Frame(container, bg='white', width=950)
        question.columnconfigure(0, weight=1)
        question.columnconfigure(1, weight=1)

        # Question title
        question_title = Label(question, text=quizzes_list[chosen_quiz][random_numbers_list[i]]['question'], font=("Arial", 15), bg = "white")
        question_title.grid(column=0, sticky="ns")
        question.pack(anchor=NW, padx=10, pady=20, expand=True, fill=BOTH)


        # RadioButtons
        for j in range(4):
            question_options = Radiobutton(question, text=quizzes_list[chosen_quiz][random_numbers_list[i]]['options'][j], bg="white",
                                        variable=intvar_list[i],
                                        value= y)
            question_options.grid(column=1, sticky=NSEW)
            y+=1 ## Chioce correctness indicator

    #Submit button
    submit = Button(container, text="submit", font=("Arial" ,10), command= submit_exam)
    submit.pack(anchor=CENTER)
    Take_Exam.mainloop()

###################### Choose Exam Page ######################

def choose_Exam():
    # close main window or the results if open
    try:
        main_window.destroy()
    except:
        1
    try:
        results_window.destroy()
    except:
        1
        
    ####### question_chooser
    
    def quizData_window(quiz_number):
        global chosen_quiz
        chosen_quiz = quiz_number-1
        if len(quiz_scores['scores'][chosen_quiz]) > 2:
            messagebox.showerror('Error', "You reached maximum number of attempts")
            return
        quizzes_window.destroy()
        TakeExam()

    quizzes_window = Tk()#22
    quizzes_window.geometry('1920x1080')
    quizzes_window.configure(background='#003154',)
    
    ##TITLE##
    title_label = Label(
        quizzes_window,
        text="Choose the quiz you want to take or edit",
        font=("Helvetica", 50, "bold italic"),
        fg="white",
        bg="#003154",
    )
    title_label.pack(side="top", pady=0)
    ##################################################
    main_frame = Frame(quizzes_window, width=1200, height=600, background='#003154')
    main_frame.pack(side="left", padx=0, pady=10, anchor="w")
    ##########images########
    image1=Image.open(r"Images\quiz.png")
    image_resize=image1.resize((400, 225))
    image_tk=ImageTk.PhotoImage(image_resize)
    image2=Image.open(r"Images\rb_85669.png")
    image_resize2=image2.resize((550, 680))
    image_tk2=ImageTk.PhotoImage(image_resize2)
    #####buttons#######
    quiz1=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="1",
                compound="center",activebackground="white", command=lambda: quizData_window(1))
    quiz1.grid(row=0, column=0, padx=10, pady=10)
    quiz2=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="2",
                compound="center",activebackground="white", command=lambda: quizData_window(2))
    quiz2.grid(row=0, column=1, padx=10, pady=10)
    quiz3=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="3",
                compound="center",activebackground="white", command=lambda: quizData_window(3))
    quiz3.grid(row=0, column=2, padx=10, pady=10)
    quiz4=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="4",
                compound="center",activebackground="white", command=lambda: quizData_window(4))
    quiz4.grid(row=1, column=0, padx=10, pady=10)
    quiz5=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="5",
                compound="center",activebackground="white", command=lambda: quizData_window(5))
    quiz5.grid(row=1, column=1, padx=10, pady=10)
    quiz6=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="6",
                compound="center",activebackground="white", command=lambda: quizData_window(6))
    quiz6.grid(row=1, column=2, padx=10, pady=10)
    quiz7=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="7",
                compound="center",activebackground="white", command=lambda: quizData_window(7))
    quiz7.grid(row=2, column=0, padx=10, pady=10)
    quiz8=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="8",
                compound="center",activebackground="white", command=lambda: quizData_window(8))
    quiz8.grid(row=2, column=1, padx=10, pady=10)
    quiz9=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="9",
                compound="center",activebackground="white", command=lambda: quizData_window(9))
    quiz9.grid(row=2, column=2, padx=10, pady=10)


    image_sd = Label(quizzes_window, image=image_tk2, background='#003154')
    image_sd.pack(side="left", padx=25, pady=0)

    quizzes_window.mainloop()
######################################


#########################################################################

####################### Create Exam #############################
def create():
    main_window.withdraw()
    def quiz(quiz_number):
        global window
        window = Toplevel()
        quizzes_window.withdraw()
        window.title("Create Exam")
        window.geometry("1920x1080")
        window.configure(bg="#003054")
        questions =[]

        Filename = 'JSON_files/quiz'+str(quiz_number)+'_questions.json'
        global load_questions
        def load_questions():
            nonlocal questions  # Access the local variable in quiz()
            try:
                with open(Filename, "r") as file:
                    questions = json.load(file)
                preview()
            except FileNotFoundError:
                pass

        # add questions to preview
        
        def preview():
            preview_listbox.delete(0, END)
            for item in questions:
                    preview_listbox.insert(END, item["question"])

        def add_question():
                question = question_entry.get()
                option1 = option1_entry.get()
                option2 = option2_entry.get()
                option3 = option3_entry.get()
                option4 = option4_entry.get()
                correct_option = correct_option_combobox.get()

                if not question or not option1 or not option2 or not option3 or not option4 or not correct_option:
                    messagebox.showwarning("Input Error", "All fields are required!")
                    return

                questions.append({
                    "question": question,
                    "options": [option1, option2, option3, option4],
                    "answer": int(correct_option)
                }
                )
                preview()

                # Clear fields for the next question
                clear()

                messagebox.showinfo("Success", "Question added successfully!")


        # Function to clear the input fields
        def clear():
            question_entry.delete(0, END)
            option1_entry.delete(0, END)
            option2_entry.delete(0, END)
            option3_entry.delete(0, END)
            option4_entry.delete(0, END)
            correct_option_combobox.set("")   # clears the correct option choice from the combobox

        # save all questions to a json file
        def save_questions():
            if not questions:
                messagebox.showwarning("Save Error", "No questions to save!")
                return
            with open(Filename, "w") as file:
                json.dump(questions, file, indent=4)
            messagebox.showinfo("Success", "Questions saved successfully!")

        # delete a selected question
        def delete_question():
            select = preview_listbox.curselection() # returns a tuple
            if not select:
                messagebox.showwarning("Delete Error", "Please select a question to delete!")
                return

            # Get the index of the selected question
            question_index = select[0]

            # Remove the question from the list
            questions.pop(question_index)

            # Update the preview
            preview()
            messagebox.showinfo("Success", "Question deleted successfully!")

        def edit_question():
            selected_index = preview_listbox.curselection()
            if not selected_index:
                messagebox.showwarning("Edit Error", "Please select a question to edit!")
                return

            index = selected_index[0]
            selected_question = questions[index]

            question_entry.delete(0, END)
            question_entry.insert(0, selected_question["question"])

            option1_entry.delete(0, END)
            option1_entry.insert(0, selected_question["options"][0])

            option2_entry.delete(0, END)
            option2_entry.insert(0, selected_question["options"][1])

            option3_entry.delete(0, END)
            option3_entry.insert(0, selected_question["options"][2])

            option4_entry.delete(0, END)
            option4_entry.insert(0, selected_question["options"][3])

            correct_option_combobox.set(str(selected_question["answer"]))

        def back():
            quizzes_window.deiconify()

        title_frame= Frame(window, bg="#006699")
        title_frame.pack(fill="x")

        title_label = Label(title_frame, text="Create Exam", font=("Arial", 30, "bold"), bg="#006699", fg="white")
        title_label.pack(pady=10)


        global add_button, edit_button, delete_button,save_button, back_button, button_frame, frame2
        frame2 = Frame(window, bg="#006699")
        frame2.pack(pady=20, fill="x")

        # Question input section
        question_label = Label(frame2 , text="Question:", font=("Arial", 20, "bold"), bg="#006699", fg="white")
        question_label.pack(anchor=W, padx=20, pady=5)
        question_entry = Entry(frame2, width=50, font=("Arial", 18))
        question_entry.pack()

        # Option inputs
        option1_label = Label(frame2, text="Option 1:", font=("Arial", 18), bg="#006699", fg="white")
        option1_label.pack(anchor=W, padx=20, pady=2)
        option1_entry = Entry(frame2, width=30, font=("Arial", 16))
        option1_entry.pack()

        option2_label = Label(frame2, text="Option 2:", font=("Arial",18), bg="#006699", fg="white")
        option2_label.pack(anchor=W, padx=20, pady=2)
        option2_entry = Entry(frame2, width=30, font=("Arial", 16))
        option2_entry.pack()

        option3_label = Label(frame2, text="Option 3:", font=("Arial", 18), bg="#006699", fg="white")
        option3_label.pack(anchor=W, padx=20, pady=2)
        option3_entry = Entry(frame2, width=30, font=("Arial", 16))
        option3_entry.pack()

        option4_label = Label(frame2, text="Option 4:", font=("Arial", 18), bg="#006699", fg="white")
        option4_label.pack(anchor=W, padx=20, pady=2)
        option4_entry = Entry(frame2, width=30, font=("Arial", 16))
        option4_entry.pack()

        # Correct option input
        correct_option_label = Label(frame2, text="Correct Option (1-4):", font=("Arial", 18), bg="#006699", fg="white")
        correct_option_label.pack(anchor=W, padx=20, pady=5)
        correct_option_combobox = ttk.Combobox(frame2, values=["1", "2", "3", "4"], font=("Arial", 16), state="readonly")
        correct_option_combobox.pack()

        # buttons
        button_frame = Frame(frame2, bg="#006699")
        button_frame.pack(pady=20)


        add_button = Button(button_frame, text="Add Question", font=("Arial", 14, "bold"), bg="#2E7D32", fg="white", command=add_question)
        add_button.grid(row=0, column=0, padx=20)

        edit_button = Button(button_frame, text="Edit Question", font=("Arial", 14, "bold"), bg="#1565C0", fg="white", command=edit_question)
        edit_button.grid(row=0, column=1, padx=20)

        delete_button = Button(button_frame, text="Delete Question", font=("Arial", 14, "bold"), bg="#D32F2F", fg="white", command=delete_question)
        delete_button.grid(row=0, column=2, padx=20)

        save_button = Button(button_frame, text="Save Quiz", font=("Arial", 14, "bold"), bg="#FBC02D", fg="white", command=save_questions)
        save_button.grid(row=0, column=3, padx=20)

        back_button = Button(button_frame, text='Back', font=("Arial", 14, "bold"), bg="black", fg="white", command=back)
        back_button.grid(row = 0, column = 4, padx=20)
        # preview questions

        preview_label = Label(window, text="Questions Preview:", font=("Arial", 18), bg="#003054", fg="white")
        preview_label.pack(anchor=W, padx=20, pady=5)

        preview_listbox = Listbox(window, width=60, height=5, font=("Arial", 16))
        preview_listbox.pack()

        window.after(100, load_questions)
        preview()

    def Quiz(quiz_number):
        quiz1_data()
        quiz(quiz_number)

    def quizData_window(quiz_number):
        quizdata_window = Toplevel()
        quizdata_window.title("Exam Creator")
        quizdata_window.configure(bg="#003054")
        filename = 'JSON_files/quiz'+str(quiz_number)+"_data.json"
        global load
        def load():
                with open(filename, "r") as file:
                    quizdata = json.load(file)
                entry1.insert(0, quizdata["Number of Questions"])
                entry2.insert(0, quizdata["Grade per Question"])
        global quiz1_data
        def quiz1_data():
            questions_number = entry1.get()
            grade = entry2.get()
            if not questions_number or not grade:
                messagebox.showwarning("Input Error","All fields are required!")
            quiz_details = {
                "Number of Questions":int(questions_number),
                "Grade per Question": int(grade)
            }
            with open(filename, 'w') as file:
                json.dump(quiz_details, file, indent=4)


        questions_label = Label(quizdata_window, text="Number of Questions:", font=('Arial', 16, 'bold'), bg="#003054", fg="white")
        questions_label.grid(row=0, column=0,sticky="N", pady=10)
        global entry1
        entry1 = Entry(quizdata_window, font=('Arial', 16, 'bold'))
        entry1.grid(row=1, column=0, padx=100)

        grade_label = Label(quizdata_window, text="Grade per Question:", font=('Arial', 16, 'bold'), bg="#003054", fg="white")
        grade_label.grid(row=2, column=0, sticky="N",pady=10)
        global entry2
        entry2 = Entry(quizdata_window, font=('Arial', 16, 'bold'))
        entry2.grid(row=2, column=0, padx=100, pady=50)

        go_to_quiz_button = Button(quizdata_window, font=('Arial', 16, 'bold'), text='Next', command=lambda: Quiz(quiz_number))
        go_to_quiz_button.grid(row=3, column=0, pady=20)
        load()
        
    #########################################################################
    
    global quizzes_window, image_tk, image_tk2
    quizzes_window = Toplevel()
    quizzes_window.geometry('1920x1080')
    quizzes_window.configure(background='#003154',)
    ##TITLE##
    title_label = Label(
        quizzes_window,
        text="Choose the quiz you want to take or edit",
        font=("Helvetica", 50, "bold italic"),
        fg="white",
        bg="#003154",
    )
    title_label.pack(side="top", pady=0)
    ##################################################
    main_frame = Frame(quizzes_window, width=1200, height=600, background='#003154')
    main_frame.pack(side="left", padx=0, pady=10, anchor="w")
    ##########images########
    image1=Image.open(r"Images\quiz.png")
    image_resize=image1.resize((400, 225))
    image_tk=ImageTk.PhotoImage(image_resize)
    image2=Image.open(r"Images\rb_85669.png")
    image_resize2=image2.resize((550, 680))
    image_tk2=ImageTk.PhotoImage(image_resize2)
    #####buttons#######
    quiz1=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="1",
                compound="center",activebackground="white", command=lambda: quizData_window(1))
    quiz1.grid(row=0, column=0, padx=10, pady=10)
    quiz2=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="2",
                compound="center",activebackground="white", command=lambda: quizData_window(2))
    quiz2.grid(row=0, column=1, padx=10, pady=10)
    quiz3=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="3",
                compound="center",activebackground="white", command=lambda: quizData_window(3))
    quiz3.grid(row=0, column=2, padx=10, pady=10)
    quiz4=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="4",
                compound="center",activebackground="white", command=lambda: quizData_window(4))
    quiz4.grid(row=1, column=0, padx=10, pady=10)
    quiz5=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="5",
                compound="center",activebackground="white", command=lambda: quizData_window(5))
    quiz5.grid(row=1, column=1, padx=10, pady=10)
    quiz6=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="6",
                compound="center",activebackground="white", command=lambda: quizData_window(6))
    quiz6.grid(row=1, column=2, padx=10, pady=10)
    quiz7=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="7",
                compound="center",activebackground="white", command=lambda: quizData_window(7))
    quiz7.grid(row=2, column=0, padx=10, pady=10)
    quiz8=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="8",
                compound="center",activebackground="white", command=lambda: quizData_window(8))
    quiz8.grid(row=2, column=1, padx=10, pady=10)
    quiz9=Button(main_frame,image=image_tk,width=400,height=225,background='#003154',font=("Helvetica", 140, "italic","bold"),fg="#9a1b00",bd=0,text="9",
                compound="center",activebackground="white", command=lambda: quizData_window(9))
    quiz9.grid(row=2, column=2, padx=10, pady=10)


    image_sd = Label(quizzes_window, image=image_tk2, background='#003154')
    image_sd.pack(side="left", padx=25, pady=0)

    ####################################################################################################
    quizzes_window.mainloop()



############################## Home Page ##########################
#window layout
def home_page():
    main_window=Toplevel()#22
    main_window.geometry('1920x1080')
    main_window.title('demo')
    main_window.configure(background='#003154')
    #media
    photo1=Image.open('Images/boy-8702235_1920.png')
    resized1=photo1.resize((400,900))
    new_photo1=ImageTk.PhotoImage(resized1)
    photo_title_beta=Image.open('Images/title.png')
    photo_title_alpha=photo_title_beta.resize((900,400))
    photo_title=ImageTk.PhotoImage(photo_title_alpha)
        #photos for buttons
    button1=PhotoImage(file='Images/take exam.png')
    button2=PhotoImage(file='Images/make exam1.png')
    button3=PhotoImage(file='Images/exit (1).png')
    button4_a=Image.open('Images/rank.png')
    button4_b=button4_a.resize((400,150))
    button4=ImageTk.PhotoImage(button4_b)

    ######################labels#########################
        #title
    title_l=Label(master=main_window,
                text="welcome to our exam creator".upper(),
                background='#003154',
                font=" calibre 40 bold",
                fg="white",
                image=photo_title,
                compound="center"
                )
    title_l.place(x=620)
    #entery and button
    parent1=Frame(main_window,
                background="#003154",
                width=1000,
                height=800
                )
    parent1.place(x=700,y=430)
    
    #### buttons 
    
    b1 = Button(parent1, text="".title(), width=380,height=330,bg="#003154", image=button1,bd=0,compound="center", font="Helvetica 30 bold",fg="#003154", command= create)
    b1.grid(row=0, column=1, padx=10, pady=10)

    b2 = Button(parent1, text="".title(), width=380,height=330,bg="#003154", image=button2,bd=0,compound="center", font="Helvetica 30 bold",fg="#088000", command= choose_Exam)
    b2.grid(row=0, column=2, padx=10, pady=10)

    b3 = Button(parent1, text="".title(), width=380,height=180,bg="#003154",image=button4,bd=0,compound="center", font="Helvetica 20 bold",fg="#088000", command=Result_page)
    b3.grid(row=2, column=1, padx=10, pady=10)

    b4 = Button(parent1, width=380,height=180,bg="#003154",image=button3,bd=0,compound="center", font="Helvetica 30 bold",fg="#088000")
    b4.grid(row=2, column=2, padx=10, pady=10)

    #student pic
    student=label = Label(main_window, image=new_photo1, background='#003154')
    student.pack(padx=50,pady=10)
    student.place(x=20,y=60)

    main_window.mainloop()


################## Create Account #######################
def create_account():
    window = Toplevel()
    window.title("Student Registration Form")
    window.geometry("1920x1080")
    window.configure(background="#003154")
    
    #### submit form
    
    def submit_form():
        name = name_entry.get().strip()
        password = password_entry.get().strip()
        email = email_entry.get().strip()
        phone = phone_entry.get().strip()
        address = address_entry.get().strip()
        age = age_entry.get().strip()

        if not name or not password or not email or not phone or not address or not age:
            messagebox.showerror("Error", "All fields are required.")
            return

        if len(password) < 6:
            messagebox.showerror("Error", "Password must be at least 6 characters.")
            return

        if "@" not in email or "." not in email:
            messagebox.showerror("Error", "Invalid email address.")
            return

        if not phone.isdigit() or len(phone) != 10:
            messagebox.showerror("Error", "Phone number must be 10 digits.")
            return

        if not age.isdigit() or int(age) <= 0:
            messagebox.showerror("Error", "Age must be a positive number.")
            return

        xlsx_path = "database.xlsx"
        try:
            workbook = load_workbook(xlsx_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
        worksheet = workbook.active

        for row in worksheet.iter_rows(values_only=True):
            if row[0] == name or row[2] == email or row[3] == phone:
                messagebox.showerror("Error", "User already exists.")
                return
        worksheet.append([name, password, email, phone, address, age])
        workbook.save(xlsx_path)

        messagebox.showinfo("Success", "Registration Successful!")
        clear_form()
        
    def clear_form():
        name_entry.delete(0, END)
        password_entry.delete(0, END)
        email_entry.delete(0, END)
        phone_entry.delete(0, END)
        address_entry.delete(0, END)
        age_entry.delete(0, END)

    def forget_password():
        x = messagebox.askquestion("Forget Password", "Are you sure you want to retrieve your password?")
        if x == "yes":
            name = name_entry.get().strip()
            email = email_entry.get().strip()
            phone = phone_entry.get().strip()
            if not name or not email or not phone:
                messagebox.showerror("Error", "Please fill in Name, Email, and Phone Number.")
                return
            try:
                workbook = load_workbook("database.xlsx")
                worksheet = workbook.active
                for row in worksheet.iter_rows(values_only=True):
                    if row[0] == name and row[2] == email and row[3] == phone:
                        password = row[1]
                        messagebox.showinfo("Password Found", f"Your password is: {password}")
                        return
                messagebox.showerror("Error", "User not found. Please check your details.")
            except Exception as e:
                messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

    def on_close():
        main_window.deiconify()
        window.destroy()

    window.protocol("WM_DELETE_WINDOW", on_close)

    try:
        image1 = Image.open('Images/create_acc.png')
        image_resize = image1.resize((700, 700))
        image_tk = ImageTk.PhotoImage(image_resize)

        image_sub = Image.open('Images/sign-up.png')
        image_resize = image_sub.resize((110, 120))
        image_submit = ImageTk.PhotoImage(image_resize)

        image_for = Image.open('Images/forgot-password.png')
        image_resize2 = image_for.resize((110, 120))
        image_forget = ImageTk.PhotoImage(image_resize2)
    except Exception as e:
        messagebox.showerror("Image Error", f"Error loading images: {e}")
        window.destroy()
        main_window.deiconify()
        return

    main_pic = Label(window, image=image_tk, background="#003154")
    main_pic.image = image_tk
    main_pic.grid(row=0, column=0, padx=150, pady=100)

    main_frame = Frame(window, bg="#003154")
    main_frame.grid(row=0, column=2)

    name_label = Label(main_frame, text="Username:", font=("Helvetica", 30, "italic", "bold"),
                       fg="#db2b00", background="#003154")
    name_label.grid(row=0, column=0, padx=10, pady=10)
    name_entry = Entry(main_frame, width=80)
    name_entry.grid(row=0, column=1, padx=10, pady=10)

    password_label = Label(main_frame, font=("Helvetica", 30, "italic", "bold"),
                           fg="#db2b00", text="Password:", background="#003154")
    password_label.grid(row=1, column=0, padx=10, pady=10)
    password_entry = Entry(main_frame, show="*", width=80)
    password_entry.grid(row=1, column=1, padx=10, pady=10)

    email_label = Label(main_frame, font=("Helvetica", 30, "italic", "bold"),
                       fg="#db2b00", text="Email:", background="#003154")
    email_label.grid(row=2, column=0, padx=10, pady=10)
    email_entry = Entry(main_frame, width=80)
    email_entry.grid(row=2, column=1, padx=10, pady=10)

    phone_label = Label(main_frame, font=("Helvetica", 30, "italic", "bold"),
                       fg="#db2b00", text="Phone Number:", background="#003154")
    phone_label.grid(row=3, column=0, padx=10, pady=10)
    phone_entry = Entry(main_frame, width=80)
    phone_entry.grid(row=3, column=1, padx=10, pady=10)

    address_label = Label(main_frame, font=("Helvetica", 30, "italic", "bold"),
                          fg="#db2b00", text="Address", background="#003154")
    address_label.grid(row=4, column=0, padx=10, pady=10)
    address_entry = Entry(main_frame, width=80)
    address_entry.grid(row=4, column=1, padx=10, pady=10)

    age_label = Label(main_frame, font=("Helvetica", 30, "italic", "bold"),
                     fg="#db2b00", text="Age:", background="#003154")
    age_label.grid(row=5, column=0, padx=10, pady=10)
    age_entry = Entry(main_frame, width=80)
    age_entry.grid(row=5, column=1, padx=10, pady=10)

    submit_button = Button(main_frame, text="Submit", image=image_submit, command=submit_form,
                           background="#003154", bd=0, activebackground="#003154")
    submit_button.image = image_submit
    submit_button.grid(row=6, column=1, padx=200, pady=10, sticky="e")

    forget_button = Button(main_frame, image=image_forget, command=forget_password,
                           background="#003154", bd=0, activebackground="#003154",
                           text='Fill the form \n to know your password',
                           compound='top', fg="white",
                           font=("Arial", 10, "italic", "bold"))
    forget_button.image = image_forget
    forget_button.grid(row=6, column=1, padx=10, pady=10, sticky="w")

def user_check():
    xlsx_path = "database.xlsx"
    try:
        workbook1 = load_workbook(xlsx_path)
        worksheet = workbook1.active
    except FileNotFoundError:
        return messagebox.showerror("Error", "Database file not found.")
    except Exception as e:
        return messagebox.showerror("Error", f"An error occurred: {e}")

    for row in worksheet.iter_rows(values_only=True):
        if row[0] == username_entry.get() and row[1] == password_entry.get():
            messagebox.showinfo("Welcome", f"Welcome back {username_entry.get()}")
            flag = True
            home_page()
            return

    return messagebox.showinfo("Login Failed", "Please check your username and password.")

def close():
    main_window.withdraw()

main_window = Tk()#22
main_window.geometry("1920x1080")
main_window.title("Create Account")
main_window.configure(background="#003154")

try:
    image1 = Image.open('Images/create__.png')
    image_resize = image1.resize((700, 590))
    image_tk = ImageTk.PhotoImage(image_resize)

    image_log = Image.open('Images/login.png')
    image_resize_login = image_log.resize((110, 110))
    image_login = ImageTk.PhotoImage(image_resize_login)

    image_cre = Image.open('Images/add-user.png')
    image_resize_create = image_cre.resize((100, 100))
    image_create = ImageTk.PhotoImage(image_resize_create)
except Exception as e:
    messagebox.showerror("Image Error", f"Error loading images: {e}")
    main_window.destroy()

title = Label(main_window, text="Welcome to our Exam Creator Project",
              font=("Helvetica", 30, "italic", "bold"), fg="white", background="#003154")
title.place(x=980, y=200)

main_image = Label(main_window, background="#003154", image=image_tk, width=700, height=590)
main_image.image = image_tk
main_image.grid(row=0, column=0, padx=80, pady=180)

frame = Frame(main_window, background="#003154", width=1000, height=800)
frame.grid(row=0, column=1, padx=100, pady=10)

username_label = Label(frame, text="Username:", font=("Helvetica", 30, "italic", "bold"),
                      fg="#db2b00", background="#003154")
username_label.grid(row=0, column=1, padx=10, pady=10)
username_entry = Entry(frame, width=80)
username_entry.grid(row=0, column=2, padx=10, pady=10)

password_label = Label(frame, font=("Helvetica", 30, "italic", "bold"),
                       fg="#db2b00", text="Password:", background="#003154")
password_label.grid(row=1, column=1, padx=10, pady=10)
password_entry = Entry(frame, show="*", width=80)
password_entry.grid(row=1, column=2, padx=10, pady=10)

login_button = Button(frame, image=image_login, width=110, height=110,
                      background='#003154', font=("Helvetica", 140, "italic", "bold"),
                      fg="#9a1b00", bd=0, activebackground="white", command=user_check)
login_button.image = image_login
login_button.grid(row=2, column=2, padx=0, pady=10)

create_button = Button(frame, image=image_create, width=130, height=130,
                       background='#003154', font=("Arial", 10, "italic", "bold"),
                       fg="white", bd=0, activebackground="white",
                       text="Create Account", compound='top', command=lambda: (create_account(), close()))
create_button.image = image_create
create_button.grid(row=2, column=1, padx=0, pady=10, sticky='e')

main_window.mainloop()
